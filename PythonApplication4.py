import openpyxl
import math

MAX=10000

wb=openpyxl.Workbook()
wb.active

ws=wb.worksheets[0]

NPN=1

for row in range(2,MAX):
    cell=ws.cell(row,1)
    cell2=ws.cell(row,2)
    cell.value=row
    #cell2.value=row

    root=int(math.sqrt(row))
    cell3=ws.cell(row,3)
    #cell2.value=root

    for number in range(2,root+1):
        if row%number==0:
            cell3.value=number
            cell2.value=NPN
            #print(row,number,"割り切れた")
            break
    
    if cell3.value==None:
        cell3.value=0
        cell2.value=NPN
        NPN=row
        print(row)

wb.save("test3.xlsx")

    
"""
    if root==2:
        if row%2==0:
            cell2.value=0
        else:
            cell2.value=1
    else:
         for number in range(2,root+1):
             if row%number==0:
                cell2.value=0
                #print(row,number,"割り切れた")
                break
    
         if cell2.value!=0:
              cell2.value=1


#wb.save("test2.xlsx")
"""

from sklearn.svm import LinearSVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score
#from sklearn.model_selection import train_test_split
#from sklearn.utils.testing import all_estimators
#import warnings

y_train=[]
x_train=[]

for row in range(2,MAX-100):
    celly=ws.cell(row,3)
    y_train.append(celly.value)

for row in range(2,MAX-100):
    cellx=ws.cell(row,1)
    cellxx=ws.cell(row,2)
    x_train.append([cellx.value,cellxx.value])
#print(x)

y_test=[]
x_test=[]

for row in range(MAX-99,MAX):
    celly=ws.cell(row,3)
    y_test.append(celly.value)

for row in range(MAX-99,MAX):
    cellx=ws.cell(row,1)
    cellxx=ws.cell(row,2)
    x_test.append([cellx.value,cellxx.value])

#x_train,x_test,y_train,y_test=train_test_split(x,y,test_size=0.2,train_size=0.8,shuffle=True)

#clf=LinearSVC()
clf=KNeighborsClassifier(n_neighbors=1)
clf.fit(x_train,y_train)
"""
warnings.filterwarnings("ignore")
allAlgorithms=all_estimators(type_filter="classifier")

for (name,algorithm) in allAlgorithms:
    clf=algorithm()
"""

y_pred=clf.predict(x_test)

MAX99=MAX-99
for y_pred_ans in y_pred:
    cell_y_pred=ws.cell(MAX99,4)
    cell_y_pred.value=y_pred_ans
    MAX99+=1

wb.save("test2.xlsx")

print("正解率=",accuracy_score(y_test,y_pred))
